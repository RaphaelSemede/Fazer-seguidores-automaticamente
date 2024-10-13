import os
from winotify import Notification, audio
from datetime import date
import instaloader
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# Caminho do arquivo Excel
file_path = 'Seguidores.xlsx'

# Verifica se o arquivo existe
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
else:
    raise FileNotFoundError(f"O arquivo {file_path} não foi encontrado.")

sheet = workbook.active

# Define a data formatada
data = date.today()
data_formatada = data.strftime("%d/%m/%Y")

# Confere se já utilizei o programa no dia
for cell in sheet[1]:
    if cell.value == data_formatada:
        print("Você já fez os seguidores de hoje!")

    elif cell.value is None:
        cell.value = data_formatada
        coluna_escrita = get_column_letter(cell.column)  # Obtém a letra da coluna onde a data foi escrita
        print(f"A data de hoje ({data_formatada}) foi escrita na célula {cell.coordinate}.")
        break
else:
    # Este else é executado se nenhum break for alcançado no loop
    print("Não há células vazias na linha 1. A data não foi escrita.")

# Inicializa o Instaloader
L = instaloader.Instaloader()

username = 'dgdsfgdsfg81'
password = 'Oba022208'

# Login
try:
    L.login(username, password)
except instaloader.exceptions.BadCredentialsException:
    raise ValueError("Credenciais inválidas. Verifique seu nome de usuário e senha.")
except instaloader.exceptions.ConnectionException:
    raise ConnectionError("Falha ao estabelecer conexão com o Instagram. Verifique sua conexão com a internet.")
except Exception as e:
    raise Exception(f"Ocorreu um erro inesperado: {e}")


# Função para encontrar a próxima célula vazia em uma linha específica
def find_next_empty_cell_in_row(sheet, row, start_col):
    col = start_col
    while sheet.cell(row=row, column=col).value is not None:
        col += 1
    return col

# Percorre todas as linhas a partir da terceira linha (linha 3)
for row in range(3, sheet.max_row + 1):
    # Obtém o nome de usuário da célula na coluna A da linha atual
    usuario = sheet.cell(row=row, column=1).value

    # Se não houver nome de usuário na célula, passa para a próxima linha
    if not usuario:
        continue

    try:
        # Verifica se o nome de usuário é uma string não vazia
        if not isinstance(usuario, str) or not usuario.strip():
            print(f"Nome de usuário inválido na linha {row}.")
            continue

        # Carrega o perfil pelo nome de usuário
        profile = instaloader.Profile.from_username(L.context, usuario)

        # Converte a coluna escrita de letra para número
        coluna_escrita_num = column_index_from_string(coluna_escrita)

        # Obtém a próxima célula vazia na mesma linha para atualizar o número de seguidores
        next_col = find_next_empty_cell_in_row(sheet, row, coluna_escrita_num)

        # Atualiza o número de seguidores na próxima célula vazia
        sheet.cell(row=row, column=next_col).value = profile.followers
    except instaloader.exceptions.ProfileNotExistsException:
        print(f"O perfil '{usuario}' não foi encontrado.")
    except instaloader.exceptions.ConnectionException as e:
        print(f"Erro de conexão ao acessar a API do Instagram: {e}")
    except instaloader.exceptions.QueryReturnedBadRequestException:
        print(f"Requisição inválida para o nome de usuário '{usuario}'. Verifique se o nome de usuário está correto.")
    except Exception as e:
        print(f"Ocorreu um erro inesperado ao processar o perfil '{usuario}': {e}")

# Salva o arquivo Excel atualizado
workbook.save(file_path)

if workbook.save == True:
    notificacao = Notification(app_id='Seguidores', title='Notificação de Conclusão', 
                               msg='Acabou de Fazer os Seguidores', duration='short')
    notificacao.show()
    

print(f'O número de seguidores foi salvo com sucesso em {file_path}')